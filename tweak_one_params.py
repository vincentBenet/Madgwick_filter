import random
from random import shuffle

import openpyxl
import numpy as np
import os
import check_ahrs
import utils

excel_path = "performances_dichotomy.xlsx"

# Parameters to optimize (wide-format order = Excel left-to-right)
param_names = [
    "gain_acc",
    "gain_mag",
    "duration_filter_mag_axis",
    "duration_filter_mag_merged",
    "duration_filter_gyr",
    "n_filter_acc",
    "duration_filter_quaternions_outputs",
    "beta",
    "adaptative_beta",
    "adaptative_gain_acc",
    "adaptative_gain_mag",
]

metric_names = ["MAG error", "ACC error", "YAW cor", "STD MAG", "ENV MAG"]
result_names = ["STD AVG", "ENV AVG"]

# Param typing
MAX1_PARAMS = {"adaptative_beta", "adaptative_gain_acc", "adaptative_gain_mag"}
INTEGER_PARAMS = {"n_filter_acc"}


def sanitize_value(pname, v):
    # Enforce types and positivity
    if pname in MAX1_PARAMS:
        return min(1, v)
    if pname in INTEGER_PARAMS:
        return max(0, int(round(v)))
    return float(max(0.0, v))


def score_from_avgs(std_avg, env_avg):
    return std_avg


def wait_if_excel_open(path):
    while True:
        try:
            with open(path, "a"):
                return
        except PermissionError:
            input(f"Fichier '{path}' ouvert dans Excel. Ferme-le puis appuie sur Entrée...")


def ensure_column(name):
    global col_index, header
    if name not in col_index:
        ws.cell(row=1, column=len(header) + 1, value=name)
        header.append(name)
        col_index[name] = len(header) - 1
    return col_index[name]


best_score = float('inf')
while True:
    if os.path.exists(excel_path):
        wait_if_excel_open(excel_path)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        header = list(param_names)
        for cfg_name in check_ahrs.configs.keys():
            for m in metric_names:
                header.append(f"{m} {cfg_name}")
        header += result_names
        ws.append(header)

    header = [cell.value for cell in ws[1]]
    col_index = {name: idx for idx, name in enumerate(header)}

    existing_combos = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        combo = tuple(v if isinstance(v, float) else v
                      for v in row[:len(param_names)])
        if any(v is None for v in combo):
            continue
        existing_combos[combo] = row_idx

    std_avg_col = ensure_column("STD AVG")
    env_avg_col = ensure_column("ENV AVG")

    best_score = float('inf')
    current_best = {p: check_ahrs.params[p] for p in param_names}

    for combo, row_idx in existing_combos.items():
        std_avg_val = ws.cell(row=row_idx, column=std_avg_col + 1).value
        env_avg_val = ws.cell(row=row_idx, column=env_avg_col + 1).value
        if std_avg_val is None or env_avg_val is None:
            continue
        try:
            score = score_from_avgs(std_avg_val, env_avg_val)
        except (TypeError, ValueError):
            continue
        if score < best_score:
            best_score = score
            for i, pname in enumerate(param_names):
                current_best[pname] = combo[i]

    print("Starting optimization with best params from Excel:")
    for p, v in current_best.items():
        print(f"  {p}: {v}")
    print(f"{best_score = }")

    pname = param_names[random.randint(0, len(param_names)-1)]
    base_value = float(current_best[pname])
    if pname in INTEGER_PARAMS:
        candidates = [base_value - 1, base_value + 1]
    elif pname in MAX1_PARAMS:
        candidates = [base_value - 10 ** -i for i in range(1, 5)] + [base_value + base_value * 10 ** -i for i
                                                                               in range(1, 5)]
    else:
        candidates = [base_value - base_value * 10**-i for i in range(5)] + [base_value + base_value * 10**-i for i in range(5)]
    shuffle(candidates)

    for val in candidates:
        val = sanitize_value(pname, val)
        print(f"TRY: {pname} = {val} / {current_best[pname]}")

        params_test = dict(current_best)
        params_test[pname] = val
        combo = tuple(params_test[p] for p in param_names)
        print(f"{current_best = }")
        print(f"{params_test = }")

        ws.append([params_test[p] for p in param_names] + [None] * (len(header) - len(param_names)))
        row_idx = ws.max_row
        existing_combos[combo] = row_idx

        run_params = dict(check_ahrs.params)
        run_params.update(params_test)
        run_params["ahrs_func"] = utils.ahrs_madgwick_python_benet
        run_params["ts_mag_to_imu"] = True

        std_sum = env_sum = 0.0
        count = 0

        for cfg_name, cfg in check_ahrs.configs.items():
            metrics_cols = {m: ensure_column(f"{m} {cfg_name}") for m in metric_names}

            if ws.cell(row=row_idx, column=metrics_cols["MAG error"] + 1).value is not None:
                std_val = ws.cell(row=row_idx, column=metrics_cols["STD MAG"] + 1).value
                env_val = ws.cell(row=row_idx, column=metrics_cols["ENV MAG"] + 1).value
                if std_val is not None and env_val is not None:
                    std_sum += std_val
                    env_sum += env_val
                    count += 1
                    print(f"  {cfg_name} already calculated")
                continue

            print(f"  Running {cfg_name}...")
            results = check_ahrs.main(
                plot=False,
                path_folder=os.path.dirname(os.path.dirname(cfg["laz"])),
                path_calibration=cfg.get("calib"),
                path_imu=cfg.get("imu"),
                path_laz_LF_ENU=None,
                **run_params
            )

            env_upgrade, _, std_upgrade, _, mag_err, acc_err, yaw_cor, *_ = results

            ws.cell(row=row_idx, column=metrics_cols["MAG error"] + 1, value=float(mag_err) * 100.0)
            ws.cell(row=row_idx, column=metrics_cols["ACC error"] + 1, value=float(acc_err) * 100.0)
            ws.cell(row=row_idx, column=metrics_cols["YAW cor"] + 1,
                    value=(float(yaw_cor) * 100.0) if not np.isnan(yaw_cor) else -100.0)
            ws.cell(row=row_idx, column=metrics_cols["STD MAG"] + 1, value=float(std_upgrade))
            ws.cell(row=row_idx, column=metrics_cols["ENV MAG"] + 1, value=float(env_upgrade))

            std_sum += float(std_upgrade)
            env_sum += float(env_upgrade)
            count += 1

            wait_if_excel_open(excel_path)
            wb.save(excel_path)

        if count > 0:
            std_avg = std_sum / count
            env_avg = env_sum / count
            ws.cell(row=row_idx, column=std_avg_col + 1, value=std_avg)
            ws.cell(row=row_idx, column=env_avg_col + 1, value=env_avg)

            s = score_from_avgs(std_avg, env_avg)
            print(f"{s = }")
            print(f"{best_score = }")
            if s < best_score:
                best_score = s
                best_val = val
                print(f"  New best for {pname}: {val} (score {s:.4f})")
                current_best[pname] = best_val
                print(f"Optimized {pname} = {best_val} (score {best_score:.4f})")

            wait_if_excel_open(excel_path)
            wb.save(excel_path)

