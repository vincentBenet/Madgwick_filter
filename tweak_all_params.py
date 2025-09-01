from scipy.optimize import minimize
import openpyxl
import numpy as np
import os
import check_ahrs
import utils

excel_path = "performances_dichotomy.xlsx"

# Parameters to optimize (wide-format order = Excel left-to-right)
param_names = [
    "gain_acc",
    "gain_mag",
    "duration_filter_mag_axis",
    "duration_filter_mag_merged",
    "duration_filter_gyr",
    "n_filter_acc",
    "duration_filter_quaternions_outputs",
    "beta",
    "adaptative_beta",
    "adaptative_gain_acc",
    "adaptative_gain_mag",
]

metric_names = ["MAG error", "ACC error", "YAW cor", "STD MAG", "ENV MAG"]
result_names = ["STD AVG", "ENV AVG"]

# Param typing
MAX1_PARAMS = {"adaptative_beta", "adaptative_gain_acc", "adaptative_gain_mag"}
INTEGER_PARAMS = {"n_filter_acc"}


def sanitize_value(pname, v):
    # Enforce types and positivity
    if pname in MAX1_PARAMS:
        return min(1, v)
    if pname in INTEGER_PARAMS:
        return max(0, int(round(v)))
    return float(max(0.0, v))


def score_from_avgs(std_avg, env_avg):
    return std_avg


def wait_if_excel_open(path):
    while True:
        try:
            with open(path, "a"):
                return
        except PermissionError:
            input(f"Fichier '{path}' ouvert dans Excel. Ferme-le puis appuie sur Entrée...")


def ensure_column(name):
    global col_index, header
    if name not in col_index:
        ws.cell(row=1, column=len(header) + 1, value=name)
        header.append(name)
        col_index[name] = len(header) - 1
    return col_index[name]


best_score = float('inf')

if os.path.exists(excel_path):
    wait_if_excel_open(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    header = list(param_names)
    for cfg_name in check_ahrs.configs.keys():
        for m in metric_names:
            header.append(f"{m} {cfg_name}")
    header += result_names
    ws.append(header)

header = [cell.value for cell in ws[1]]
col_index = {name: idx for idx, name in enumerate(header)}

existing_combos = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    combo = tuple(v if isinstance(v, float) else v
                  for v in row[:len(param_names)])
    if any(v is None for v in combo):
        continue
    existing_combos[combo] = row_idx

std_avg_col = ensure_column("STD AVG")
env_avg_col = ensure_column("ENV AVG")

best_score = float('inf')
current_best = {p: check_ahrs.params[p] for p in param_names}

# Initial guess from Excel best
x0 = np.array([current_best[p] for p in param_names], dtype=float)

# Bounds for each parameter
bounds = []
for pname in param_names:
    if pname in INTEGER_PARAMS:
        bounds.append((0, None))  # integers, rounded later
    elif pname in MAX1_PARAMS:
        bounds.append((0.0, 1.0))
    else:
        bounds.append((0.0, None))  # only positive


def objective(x):
    params_test = {pname: sanitize_value(pname, val)
                   for pname, val in zip(param_names, x)}
    combo = tuple(round(params_test[p], 5) for p in param_names)

    print(f"{combo = }")

    if combo in existing_combos:
        row_idx = existing_combos[combo]
    else:
        ws.append([params_test[p] for p in param_names] +
                  [None] * (len(header) - len(param_names)))
        row_idx = ws.max_row
        existing_combos[combo] = row_idx

    run_params = dict(check_ahrs.params)
    run_params.update(params_test)
    run_params["ahrs_func"] = utils.ahrs_madgwick_python_benet
    run_params["ts_mag_to_imu"] = True

    std_sum = env_sum = 0.0
    count = 0

    # Loop over all configs
    for cfg_name, cfg in check_ahrs.configs.items():
        metrics_cols = {m: ensure_column(f"{m} {cfg_name}") for m in metric_names}

        # Check if this config is already computed
        std_val = ws.cell(row=row_idx, column=metrics_cols["STD MAG"] + 1).value
        env_val = ws.cell(row=row_idx, column=metrics_cols["ENV MAG"] + 1).value

        if std_val is not None and env_val is not None:
            # Reuse stored values
            std_sum += float(std_val)
            env_sum += float(env_val)
            count += 1
            print(f"Reusing {cfg_name} for {params_test}")
            continue

        # Missing this config → run computation
        print(f"Running {cfg_name}...")
        results = check_ahrs.main(
            plot=False,
            path_folder=os.path.dirname(os.path.dirname(cfg["laz"])),
            path_calibration=cfg.get("calib"),
            path_imu=cfg.get("imu"),
            path_laz_LF_ENU=None,
            export="",
            **run_params
        )

        env_upgrade, _, std_upgrade, _, mag_err, acc_err, yaw_cor, *_ = results

        ws.cell(row=row_idx, column=metrics_cols["MAG error"] + 1, value=float(mag_err) * 100.0)
        ws.cell(row=row_idx, column=metrics_cols["ACC error"] + 1, value=float(acc_err) * 100.0)
        ws.cell(row=row_idx, column=metrics_cols["YAW cor"] + 1,
                value=(float(yaw_cor) * 100.0) if not np.isnan(yaw_cor) else -100.0)
        ws.cell(row=row_idx, column=metrics_cols["STD MAG"] + 1, value=float(std_upgrade))
        ws.cell(row=row_idx, column=metrics_cols["ENV MAG"] + 1, value=float(env_upgrade))

        std_sum += float(std_upgrade)
        env_sum += float(env_upgrade)
        count += 1

    # Compute averages and score
    if count > 0:
        std_avg = std_sum / count
        env_avg = env_sum / count
        ws.cell(row=row_idx, column=std_avg_col + 1, value=std_avg)
        ws.cell(row=row_idx, column=env_avg_col + 1, value=env_avg)
        score = score_from_avgs(std_avg, env_avg)
    else:
        score = 1e9  # penalty if nothing ran

    wait_if_excel_open(excel_path)
    wb.save(excel_path)

    print(f"Tested params: {params_test} \n\t-> score {score}")
    return score


res = minimize(objective, x0, method="Nelder-Mead", bounds=bounds, options={"maxiter": 1000, "disp": True})

print("Optimization complete!")
print("Best params:", dict(zip(param_names, res.x)))
print("Best score:", res.fun)
