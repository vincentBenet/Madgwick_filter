import itertools

import numpy
import openpyxl
import check_ahrs
import utils
import os
from tqdm import tqdm

excel_path = "performances.xlsx"

param_names = [
    "python_madgwick_benet",
    "ts_mag_to_imu",
    "filter_mag_axis",
    "filter_mag_merged",
    "filter_gyr",
    "filter_acc",
    "filter_quaternions",
    "beta",
    "adaptative_beta",
    "adaptative_acc",
    "adaptative_mag"
]

# Nouvelles métriques par acquisition
metric_names = ["MAG error", "ACC error", "YAW cor", "STD MAG", "ENV MAG"]
result_names = ["STD AVG", "ENV AVG"]


def wait_if_excel_open(path):
    while True:
        try:
            with open(path, "a"):
                return
        except PermissionError:
            input(f"Fichier '{path}' ouvert dans Excel. Ferme-le puis appuie sur Entrée...")


# Chargement ou création du fichier Excel
if os.path.exists(excel_path):
    wait_if_excel_open(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    header = list(param_names)
    for cfg_name in check_ahrs.configs.keys():
        for m in metric_names:
            header.append(f"{m} {cfg_name}")
    header += result_names
    ws.append(header)

# Gestion dynamique des colonnes
header = [cell.value for cell in ws[1]]
col_index = {name: idx for idx, name in enumerate(header)}


def ensure_column(name):
    global col_index, header
    if name not in col_index:
        ws.cell(row=1, column=len(header) + 1, value=name)
        header.append(name)
        col_index[name] = len(header) - 1
    return col_index[name]


existing_combos = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    combo = tuple(row[:len(param_names)])
    if all(v is not None for v in combo):
        existing_combos[combo] = row_idx

all_combos = list(itertools.product([0, 1], repeat=len(param_names)))
pbar = tqdm(total=len(all_combos), desc="Calculs", unit="combos")

for combo in all_combos:
    print(f"{combo = }")
    if (
        (sum(combo) < len(combo) - 2) or
        (not (combo[0]) and (combo[7] or combo[8] or combo[9] or combo[10])) or
        (not (combo[1]))
    ):
        print(f"Skipping {combo}")
        continue

    pbar.update(1)

    # Ligne existante ou nouvelle
    if combo in existing_combos:
        row_idx = existing_combos[combo]
    else:
        ws.append(list(combo) + [None] * (len(header) - len(param_names)))
        row_idx = ws.max_row
        existing_combos[combo] = row_idx

    print(f"{row_idx = }")

    params_test = {
        "ahrs_func": utils.ahrs_madgwick_python_benet if combo[0] else utils.ahrs_madgwick_rust,
        "ts_mag_to_imu": bool(combo[1]),
        "duration_filter_mag_axis": check_ahrs.params["duration_filter_mag_axis"] if combo[2] else 0,
        "duration_filter_mag_merged": check_ahrs.params["duration_filter_mag_merged"] if combo[3] else 0,
        "duration_filter_gyr": check_ahrs.params["duration_filter_gyr"] if combo[4] else 0,
        "n_filter_acc": check_ahrs.params["n_filter_acc"] if combo[5] else 0,
        "duration_filter_quaternions_outputs": check_ahrs.params["duration_filter_quaternions_outputs"] if combo[6] else 0,
        "beta": check_ahrs.params["beta"] if combo[7] else 1,
        "adaptative_beta": check_ahrs.params["adaptative_beta"] if combo[8] else 0,
        "adaptative_gain_acc": check_ahrs.params["adaptative_gain_acc"] if combo[9] else 0,
        "adaptative_gain_mag": check_ahrs.params["adaptative_gain_mag"] if combo[10] else 0
    }

    std_sum, env_sum, count = 0, 0, 0

    for cfg_name, cfg in check_ahrs.configs.items():
        metrics_cols = {m: ensure_column(f"{m} {cfg_name}") for m in metric_names}

        # Vérifier si déjà calculé
        if ws.cell(row=row_idx, column=metrics_cols["MAG error"] + 1).value is not None:
            std_sum += ws.cell(row=row_idx, column=metrics_cols["STD MAG"] + 1).value or 0
            env_sum += ws.cell(row=row_idx, column=metrics_cols["ENV MAG"] + 1).value or 0
            count += 1
            print(f"{cfg_name} already calculated with params {combo}")
            continue

        results = check_ahrs.main(
            plot=False,
            path_folder=os.path.dirname(os.path.dirname(cfg["laz"])),
            path_calibration=cfg.get("calib"),
            path_imu=cfg.get("imu"),
            path_laz_LF_ENU=cfg["laz"],
            gain_acc=check_ahrs.params["gain_acc"],
            gain_mag=check_ahrs.params["gain_mag"],
            export="",
            **params_test
        )
        print(params_test)

        env_upgrade, env_orion, std_upgrade, std_orion, mag_err, acc_err, yaw_cor, *_ = results

        ws.cell(row=row_idx, column=metrics_cols["MAG error"] + 1, value=mag_err * 100)
        ws.cell(row=row_idx, column=metrics_cols["ACC error"] + 1, value=acc_err * 100)
        ws.cell(row=row_idx, column=metrics_cols["YAW cor"] + 1, value=yaw_cor * 100 if not numpy.isnan(yaw_cor) else -100)
        ws.cell(row=row_idx, column=metrics_cols["STD MAG"] + 1, value=std_upgrade)
        ws.cell(row=row_idx, column=metrics_cols["ENV MAG"] + 1, value=env_upgrade)
        std_sum += std_upgrade
        env_sum += env_upgrade
        count += 1
        wait_if_excel_open(excel_path)
        wb.save(excel_path)

    # Mise à jour des moyennes
    if count > 0:
        std_avg_col = ensure_column("STD AVG")
        env_avg_col = ensure_column("ENV AVG")
        ws.cell(row=row_idx, column=std_avg_col + 1, value=std_sum / count)
        ws.cell(row=row_idx, column=env_avg_col + 1, value=env_sum / count)
        wait_if_excel_open(excel_path)
        wb.save(excel_path)

pbar.close()
